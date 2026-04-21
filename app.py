from flask import Flask, request, send_file, render_template, jsonify
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import io, os

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

NETWORK_COLORS = {
    'zong':    '#d62b2b',
    'ufone':   '#00a651',
    'jazz':    '#f7941d',
    'telenor': '#0073c2'
}
NETWORK_NAMES = {
    'zong':'Zong', 'ufone':'Ufone', 'jazz':'Jazz', 'telenor':'Telenor'
}

SERVICE_NUMBERS = {
    'zong': {
        "230","6009","310","1700","15","1122","6008","25","102","211",
        "700","777","828","829","2161","2300","3238","3239","3458","3557",
        "3737","6911","7028","7078","7258","7861","8227","8558","9080",
        "44342","47650","2545"
    },
    'ufone': {
        "blank","414b5548","4554484e4943","476f50","497466617120486f6d657",
        "5054434c","536869666120496e742e","55464f4e45","55666f6e65",
        "180","1166","3404","6525","55506169","1219","INTERNET","UNKNOWN"
    },
    'jazz': {
        "(blank)","3111","5188","8696","33313131","80000016","302771212",
        "2353494D4C4147","4A415A5A","4A415A5A3447","4A617A7A","4A617A7A203447",
        "4A617A7A436173","2211","2299","3737","8388","32323131","4A617A7A74756E",
        "111","5716","8300","123","668","3444","3445","6009","6060","6064",
        "6080","6633","7770","1","5","8","47","70","188","558","773","940",
        "3977","6381","51876","347534","8885988","MO","ikTok","hatsapp",
        "BAUTH","azzCash","azz 4G","azz","4A415A5A203447","5.34555E+57","424"
    },
    'telenor': {
        "(blank)","230","5797","6557","7770","7788","8632","727251","727287",
        "150MB43Mins33","Bari Bachat","BudgetOffer","CALLING0Rs6","Din Bhar",
        "Freefire","FULLDAYCALL","internet","MUFT 3GB","PakvAusx","Telenorx",
        "WhtsApp0Rs5","Win Balance","3737","Eid Special"
    }
}

CARRIER_PREFIXES = {
    'zong':    ["110", "92", "38"],
    'ufone':   ["9292", "92"],
    'jazz':    ["640", "92", "38", "40", "2"],
    'telenor': ["9292", "92"]
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def thin_border():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)

def autofit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except: pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

def clean(val):
    if val is None: return ''
    import math
    if isinstance(val, float) and math.isnan(val): return ''
    s = str(val).strip()
    return '' if s.lower() in ('nan','none') else s

def strip_prefix(value, network):
    s = clean(value)
    if not s or s == '---': return s
    for prefix in CARRIER_PREFIXES.get(network, []):
        if s.startswith(prefix):
            return s[len(prefix):]
    return s

def read_file(file_bytes, filename):
    ext = filename.rsplit('.', 1)[-1].lower()
    if ext == 'csv':
        for enc in ['utf-8', 'latin-1', 'cp1252']:
            try:
                return pd.read_csv(io.BytesIO(file_bytes), header=None, dtype=str, encoding=enc)
            except: continue
        raise ValueError("Cannot read CSV")
    elif ext in ['xlsx', 'xls']:
        engine = 'openpyxl' if ext == 'xlsx' else 'xlrd'
        return pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str, engine=engine)
    raise ValueError(f"Unsupported: .{ext}")

# ── ZONG processor ────────────────────────────────────────────────────────────

def process_zong(df):
    num  = clean(df.iloc[0, 1]) if df.shape[1] > 1 else ''
    name = clean(df.iloc[1, 1]) if df.shape[0] > 1 and df.shape[1] > 1 else ''
    cnic = clean(df.iloc[2, 1]) if df.shape[0] > 2 and df.shape[1] > 1 else ''
    info_rows = pd.DataFrame([['Number', num], ['Name', name], ['CNIC', cnic]])

    # Find header row containing CALL_TYPE
    hdr_idx = None
    for i in range(len(df)):
        row_vals = [clean(v).upper() for v in df.iloc[i]]
        if 'CALL_TYPE' in row_vals:
            hdr_idx = i
            break
    if hdr_idx is None:
        raise ValueError("Cannot find header row in Zong file")

    raw_hdrs = list(df.iloc[hdr_idx])
    n = len(raw_hdrs)
    # Move col D (BNUMBER) before col C (STRT_TM)
    new_order = [0, 1, 3, 2] + list(range(4, n)) if n >= 4 else list(range(n))

    headers = []
    for idx in new_order:
        h = clean(raw_hdrs[idx]) if idx < n else f'Col{idx}'
        headers.append(h)

    if len(headers) > 0: headers[0] = 'Call Type'
    if len(headers) > 1: headers[1] = 'A Party'
    if len(headers) > 2: headers[2] = 'B Party'
    if len(headers) > 3: headers[3] = 'Date & Time'
    if len(headers) > 9: headers[9] = 'Location'

    data = df.iloc[hdr_idx + 1:].copy()
    if n >= 4:
        data = data.iloc[:, new_order].copy()
    data.columns = headers[:data.shape[1]]
    data = data.reset_index(drop=True)
    data = data[data.apply(lambda r: any(clean(v) for v in r), axis=1)]

    if 'B Party' in data.columns:
        data['B Party'] = data['B Party'].apply(lambda x: strip_prefix(x, 'zong'))

    return info_rows, headers[:data.shape[1]], data.reset_index(drop=True)

# ── UFONE processor ───────────────────────────────────────────────────────────

def process_ufone(df):
    sub_num = clean(df.iloc[1, 2]) if df.shape[0] > 1 and df.shape[1] > 2 else ''
    info_rows = pd.DataFrame([['Number', sub_num], ['Name', ''], ['CNIC', '']])

    raw_hdrs = [clean(v) for v in df.iloc[0]]
    rename_map = {
        'a number': 'A Party', 'b number': 'B Party',
        'start time': 'Date & Time', 'type': 'Call Type',
        'direction': 'Direction', 'duration': 'Duration',
        'location': 'Location', 'cell id': 'Cell ID',
        'latitude': 'Latitude', 'longitude': 'Longitude',
        'imei': 'IMEI', 'imsi': 'IMSI',
    }
    headers = [rename_map.get(h.lower().strip(), h) for h in raw_hdrs]

    data = df.iloc[1:].copy()
    data.columns = headers[:data.shape[1]]
    data = data.reset_index(drop=True)

    if 'B Party' in data.columns and sub_num:
        data['B Party'] = data['B Party'].replace(sub_num, '')

    data = data.fillna('---')
    for col in data.columns:
        data[col] = data[col].apply(lambda x: '---' if clean(str(x)) == '' else clean(str(x)))

    if 'B Party' in data.columns:
        data['B Party'] = data['B Party'].apply(
            lambda x: strip_prefix(x, 'ufone') if x != '---' else x)

    if 'Date & Time' in data.columns:
        try: data = data.sort_values('Date & Time').reset_index(drop=True)
        except: pass

    keep = ['A Party','B Party','Call Type','Direction','Date & Time',
            'Duration','Cell ID','IMEI','IMSI','Location','Latitude','Longitude']
    final_hdrs = [h for h in keep if h in data.columns]
    return info_rows, final_hdrs, data[final_hdrs].reset_index(drop=True)

# ── JAZZ processor ────────────────────────────────────────────────────────────

def process_jazz(df):
    sub_num = clean(df.iloc[1, 1]) if df.shape[0] > 1 and df.shape[1] > 1 else ''
    info_rows = pd.DataFrame([['Number', sub_num], ['Name', ''], ['CNIC', '']])

    raw_hdrs = [clean(v) for v in df.iloc[0]]
    rename_map = {
        'calltype': 'Call Type', 'aparty': 'A Party', 'bparty': 'B Party',
        'datetime': 'Date & Time', 'duration': 'Duration', 'cellid': 'Cell ID',
        'imsi': 'IMSI', 'imei': 'IMEI', 'sitelocation': 'Location',
    }
    headers = [rename_map.get(h.lower().strip().replace(' ','').replace('_',''), h)
               for h in raw_hdrs]

    max_col = min(9, df.shape[1])
    data = df.iloc[1:, :max_col].copy()
    data.columns = headers[:max_col]
    data = data.reset_index(drop=True)

    if 'B Party' in data.columns:
        data['B Party'] = data['B Party'].apply(lambda x: strip_prefix(x, 'jazz'))

    return info_rows, list(data.columns), data.reset_index(drop=True)

# ── TELENOR processor ─────────────────────────────────────────────────────────

def process_telenor(df):
    sub_num = clean(df.iloc[1, 0]) if df.shape[0] > 1 else ''
    info_rows = pd.DataFrame([['Number', sub_num], ['Name', ''], ['CNIC', '']])

    data = df.iloc[1:].copy()
    data.columns = [clean(v) for v in df.iloc[0]]
    data = data.reset_index(drop=True)

    start_col = next((c for c in data.columns if 'start' in c.lower()), None)
    if start_col:
        try: data = data.sort_values(start_col).reset_index(drop=True)
        except: pass

    a_col   = next((c for c in data.columns if c.lower() == 'msisdn'), None)
    b_col   = next((c for c in data.columns if 'call_org_num' in c.lower()), None)
    b2_col  = next((c for c in data.columns if 'call_dialed_num' in c.lower()), None)
    dt_col  = next((c for c in data.columns if 'start' in c.lower()), None)
    dir_col = next((c for c in data.columns if 'inbound_outbound' in c.lower()), None)
    dur_col = next((c for c in data.columns if 'network_volume' in c.lower()), None)
    ct_col  = next((c for c in data.columns if c.lower() == 'call_type'), None)
    loc_col = next((c for c in data.columns if c.lower() == 'location'), None)
    lac_col = next((c for c in data.columns if 'lac' in c.lower()), None)
    site_col= next((c for c in data.columns if 'site_id' in c.lower()), None)

    if b_col and sub_num:
        data[b_col] = data[b_col].replace(sub_num, '')
    if b_col and b2_col:
        mask = data[b_col].apply(lambda x: clean(str(x)) == '')
        data.loc[mask, b_col] = data.loc[mask, b2_col]

    if dir_col and ct_col:
        data['Call Type'] = (data[dir_col].apply(clean)+' '+data[ct_col].apply(clean)).str.strip()
    elif ct_col:
        data['Call Type'] = data[ct_col]
    else:
        data['Call Type'] = ''

    out = pd.DataFrame()
    if a_col:   out['A Party']     = data[a_col].apply(clean)
    if b_col:   out['B Party']     = data[b_col].apply(clean)
    if dt_col:  out['Date & Time'] = data[dt_col].apply(clean)
    out['Call Type'] = data['Call Type']
    if dur_col: out['Duration']    = data[dur_col].apply(clean)
    if lac_col: out['LAC']         = data[lac_col].apply(clean)
    if site_col:out['Site ID']     = data[site_col].apply(clean)
    if loc_col: out['Location']    = data[loc_col].apply(clean)

    out = out.fillna('---')
    for col in out.columns:
        out[col] = out[col].apply(lambda x: '---' if clean(str(x)) == '' else clean(str(x)))

    if 'B Party' in out.columns:
        out['B Party'] = out['B Party'].apply(
            lambda x: strip_prefix(x, 'telenor') if x != '---' else x)

    return info_rows, list(out.columns), out.reset_index(drop=True)

# ── Sheet builders ────────────────────────────────────────────────────────────

def build_cdr_sheet(ws, info_rows, headers, data_df, network):
    bold14   = Font(size=14, bold=True)
    bold     = Font(bold=True)
    hdr_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
    inf_fill = PatternFill(fill_type='solid', fgColor='FFF2CC')

    labels = ['Number','Name','CNIC']
    for ri, label in enumerate(labels, start=1):
        c = ws.cell(row=ri, column=1, value=label)
        c.font=bold14; c.border=thin_border(); c.fill=inf_fill
        ws.merge_cells(f'B{ri}:C{ri}')
        val = clean(info_rows.iloc[ri-1,1]) if ri-1 < len(info_rows) and info_rows.shape[1]>1 else ''
        vc = ws.cell(row=ri, column=2, value=val)
        vc.font=bold14; vc.border=thin_border()

    for ci, hdr in enumerate(headers, start=1):
        c = ws.cell(row=5, column=ci, value=hdr)
        c.font=bold; c.border=thin_border(); c.fill=hdr_fill

    for ri, row_data in data_df.iterrows():
        er = ri + 6
        for ci, val in enumerate(row_data, start=1):
            c = ws.cell(row=er, column=ci, value=clean(val) if pd.notna(val) else '')
            c.border=thin_border()

    autofit_columns(ws)
    ws.page_setup.orientation='landscape'; ws.page_setup.paperSize=5
    ws.page_setup.scale=70; ws.print_title_rows='5:5'
    ws.page_margins.top=0.41; ws.page_margins.right=0.25
    ws.page_margins.bottom=0.38; ws.page_margins.left=1.14

def build_summary_sheet(ws, data_df, network):
    bold     = Font(bold=True)
    hdr_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
    wht_fill = PatternFill(fill_type='solid', fgColor='FFFFFF')

    a_col  = next((c for c in data_df.columns if 'a party' in c.lower()), None)
    b_col  = next((c for c in data_df.columns if 'b party' in c.lower()), None)
    ct_col = next((c for c in data_df.columns if 'call type' in c.lower().replace('_',' ')), None)

    if not all([a_col, b_col, ct_col]):
        ws.cell(row=3,column=1,value='Summary unavailable — missing columns'); return

    svc = SERVICE_NUMBERS.get(network, set())
    df  = data_df[~data_df[b_col].astype(str).isin(svc)].copy()
    df  = df[~df[b_col].astype(str).str.strip().isin(['','---'])].copy()

    call_types = sorted(df[ct_col].dropna().unique())
    try:
        pivot = df.groupby([a_col,b_col])[ct_col].value_counts().unstack(fill_value=0)
        for ct in call_types:
            if ct not in pivot.columns: pivot[ct] = 0
        pivot = pivot[call_types]
        pivot['Grand Total'] = pivot.sum(axis=1)
        pivot = pivot.sort_values('Grand Total', ascending=False).reset_index()
    except Exception as e:
        ws.cell(row=3,column=1,value=f'Summary error: {e}'); return

    col_headers = [a_col, b_col] + call_types + ['Grand Total']
    for ci, hdr in enumerate(col_headers, start=1):
        c = ws.cell(row=3,column=ci,value=hdr)
        c.font=bold; c.border=thin_border(); c.fill=hdr_fill

    for ri, row_data in pivot.iterrows():
        er = ri + 4
        for ci, val in enumerate(row_data, start=1):
            c = ws.cell(row=er,column=ci,value=val)
            c.border=thin_border(); c.fill=wht_fill

    if a_col in pivot.columns:
        a_parties = ', '.join(
            str(v) for v in pivot[a_col].unique()
            if str(v).strip() not in ('','Grand Total','nan'))
        ws.cell(row=2, column=len(col_headers), value=a_parties)

    autofit_columns(ws)
    ws.page_setup.scale=70
    ws.page_margins.top=0.41; ws.page_margins.right=0.25
    ws.page_margins.bottom=0.38; ws.page_margins.left=1.14

def build_i2_sheet(ws, headers, data_df, network):
    bold     = Font(bold=True)
    hdr_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
    strip_col = 'B Party' if network in ('zong','jazz') else 'A Party'

    for ci, hdr in enumerate(headers, start=1):
        c = ws.cell(row=1,column=ci,value=hdr)
        c.font=bold; c.border=thin_border(); c.fill=hdr_fill

    for ri, row_data in data_df.iterrows():
        er = ri + 2
        for ci, (col_name, val) in enumerate(row_data.items(), start=1):
            s = clean(val) if pd.notna(val) else ''
            if col_name == strip_col and len(s) > 2:
                s = s[2:]
            c = ws.cell(row=er,column=ci,value=s)
            c.border=thin_border()

    autofit_columns(ws)

def create_output_excel(info_rows, headers, data_df, network):
    wb = Workbook()
    ws_cdr     = wb.active; ws_cdr.title='CDR'
    ws_summary = wb.create_sheet('Summary')
    ws_i2      = wb.create_sheet('I2')
    build_cdr_sheet(ws_cdr, info_rows, headers, data_df, network)
    build_summary_sheet(ws_summary, data_df, network)
    build_i2_sheet(ws_i2, headers, data_df, network)
    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out

# ── Flask routes ──────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    if 'file' not in request.files:
        return jsonify({'error':'No file uploaded'}), 400
    file    = request.files['file']
    network = request.form.get('network','').lower()

    if not file.filename:
        return jsonify({'error':'No file selected'}), 400
    if network not in ('zong','ufone','jazz','telenor'):
        return jsonify({'error':'Please select a network'}), 400

    ext = file.filename.rsplit('.',1)[-1].lower()
    if ext not in {'xlsx','xls','csv'}:
        return jsonify({'error':f'.{ext} not supported'}), 400

    try:
        raw    = file.read()
        df_raw = read_file(raw, file.filename)
        processors = {
            'zong':process_zong,'ufone':process_ufone,
            'jazz':process_jazz,'telenor':process_telenor
        }
        info_rows, headers, data_df = processors[network](df_raw)
        output   = create_output_excel(info_rows, headers, data_df, network)
        out_name = file.filename.rsplit('.',1)[0] + f'_{NETWORK_NAMES[network]}_Processed.xlsx'
        return send_file(output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=out_name)
    except Exception as e:
        return jsonify({'error':str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
